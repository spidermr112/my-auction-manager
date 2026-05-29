import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
import os

# 💡 연락처 하이픈 자동 변환 함수
def format_phone_number(input_str):
    if not input_str:
        return ""
    # 숫자만 추출
    nums = "".join(filter(str.isdigit, input_str))
    
    # 8자리인 경우 (예: 12345678) -> 앞에 010 붙이기
    if len(nums) == 8:
        nums = "010" + nums
        
    # 11자리 완성형인 경우 하이픈 포맷팅
    if len(nums) == 11:
        return f"{nums[:3]}-{nums[3:7]}-{nums[7:]}"
    # 10자리인 경우 예외 처리
    elif len(nums) == 10:
        return f"{nums[:3]}-{nums[3:6]}-{nums[6:]}"
        
    return input_str

class RealEstateManager:
    def __init__(self, root):
        self.root = root
        self.root.title("부동산 매니저 v5.0")
        self.root.geometry("480x850")
        self.root.configure(bg="white")
        self.label_bg = "#d9ead3"
        self.file_name = "RealEstate_Data.xlsx"
        self.inputs = {}
        self.create_widgets()
        self.update_next_id()
        
    def create_widgets(self):
        # 헤더 섹션
        header_frame = tk.Frame(self.root, bg="#ffd966", pady=10)
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="매물등록", font=("Arial", 20, "bold"), bg="#ffd966").pack()

        # 입력 필드 섹션
        input_frame = tk.Frame(self.root, bg="white", padx=10, pady=10)
        input_frame.pack(fill="both", expand=True)

        # 목록 데이터 설정
        trade_types = ["월세", "전세", "매매"]
        property_types = ["빌라", "아파트", "단독", "오피스텔", "상가", "창고", "공장", "토지"]
        room_options = ["방1", "방2", "방3", "방4", "방5"]
        bath_options = ["욕실1", "욕실2", "욕실3", "욕실4"]
        elevator_options = ["유", "무"]

        # 필드 구성
        field_configs = [
            ("매물번호", "", "entry"),
            ("접수일자", datetime.now().strftime("%Y/%m/%d"), "entry"),
            ("구      분", trade_types, "combo"),
            ("물건종류", property_types, "combo"),
            ("방      ", room_options, "combo"),
            ("욕      실", bath_options, "combo"),
            ("층/총층", "", "entry"),
            ("엘리베이터", elevator_options, "combo"),
            ("공급/전용", "", "entry"),
            ("상세정보", "", "entry"),
            ("면      적", "", "entry"),
            ("거래가액", "", "entry"),
            ("주      소", "남양주시 ", "entry"),
            ("소 유 자", "", "entry"),
            ("세 입 자", "", "entry"),
            ("연 락 처", "", "entry"),  # 💡 여기에 연락처 칸을 배치했습니다.
            ("특약사항", "", "text")
        ]

        for i, (label, val, w_type) in enumerate(field_configs):
            row, col = divmod(i, 2)
            lbl = tk.Label(input_frame, text=label, bg=self.label_bg, width=10, relief="ridge")
            lbl.grid(row=row, column=col*2, sticky="nsew", pady=2)

            if w_type == "entry":
                ent = tk.Entry(input_frame)
                ent.insert(0, val)
                ent.grid(row=row, column=col*2+1, sticky="nsew", padx=2)
                self.inputs[label] = ent
                
                # 💡 연락처 칸에 포커스가 나갈 때 자동 변환되도록 이벤트 바인딩
                if label == "연 락 처":
                    ent.bind("<FocusOut>", self.auto_format_phone)
                    
            elif w_type == "combo":
                cb = ttk.Combobox(input_frame, values=val, state="readonly")
                cb.current(0)
                cb.grid(row=row, column=col*2+1, sticky="nsew", padx=2)
                self.inputs[label] = cb
            elif w_type == "text":
                txt = tk.Text(input_frame, height=4)
                txt.grid(row=row, column=col*2+1, columnspan=3, sticky="nsew", padx=2, pady=5)
                self.inputs[label] = txt

        # 저장 버튼
        save_btn = tk.Button(self.root, text="저 장", bg="#ed9121", font=("Arial", 12, "bold"), command=self.save_to_excel, height=2)
        save_btn.pack(fill="x", padx=20, pady=10)

    # 💡 마우스 탈출 시 호출되는 자동 변환 실행 함수
    def auto_format_phone(self, event):
        widget = self.inputs["연 락 처"]
        current_text = widget.get().strip()
        formatted = format_phone_number(current_text)
        widget.delete(0, tk.END)
        widget.insert(0, formatted)

    def update_next_id(self):
        next_num = 1
        if os.path.exists(self.file_name):
            try:
                wb = openpyxl.load_workbook(self.file_name)
                ws = wb.active
                if ws.max_row > 1:
                    last_val = ws.cell(row=ws.max_row, column=1).value
                    if last_val and str(last_val).isdigit():
                        next_num = int(last_val) + 1
                wb.close()
            except Exception:
                next_num = 1
        self.inputs["매물번호"].delete(0, tk.END)
        self.inputs["매물번호"].insert(0, str(next_num).zfill(8))

    def save_to_excel(self):
        current_data = []
        headers = list(self.inputs.keys())
        for label in headers:
            widget = self.inputs[label]
            if isinstance(widget, tk.Text):
                val = widget.get("1.0", "end-1c").strip()
            else:
                val = widget.get().strip()
            current_data.append(val)

        try:
            if not os.path.exists(self.file_name):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)
            else:
                wb = openpyxl.load_workbook(self.file_name)
                ws = wb.active
            ws.append(current_data)
            wb.save(self.file_name)
            messagebox.showinfo("성공", "데이터가 엑셀에 저장되었습니다.")

            # --- 저장 후 입력창 초기화 로직 ---
            for label, widget in self.inputs.items():
                if label in ["매물번호", "접수일자"]:
                    continue
                if isinstance(widget, tk.Entry):
                    widget.delete(0, tk.END)
                    if "주      소" in label:
                        widget.insert(0, "남양주시 ")
                elif isinstance(widget, tk.Text):
                    widget.delete("1.0", tk.END)
                elif isinstance(widget, ttk.Combobox):
                    widget.current(0)
            
            self.update_next_id()
        except PermissionError:
            messagebox.showerror("오류", "엑셀 파일을 닫고 다시 시도해 주세요.")
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = RealEstateManager(root)
    root.mainloop()
